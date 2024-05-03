import pandas as pd
from tests import cp0011, cp0012, cp0021, cp0022, cp0031, cp0032, cp0041, cp0042, cp0052, cp0051, cp0061, cp0062, cp0071, cp0072, cp0081, cp0082, cp0091, cp0092, cp0101, cp0102, cp0111, cp0112, cp0122, cp0121
from openpyxl.styles import PatternFill

from generador import generate_user_data
import os
tests_names = ['cp0011', 'cp0012', 'cp0021', 'cp0022', 'cp0031', 'cp0032', 'cp0041', 'cp0042', 'cp0051', 'cp0052', 'cp0061', 'cp0062', 'cp0071', 'cp0072', 'cp0081', 'cp0082', 'cp0091', 'cp0092', 'cp0101', 'cp0102', 'cp0111', 'cp0112', 'cp0121', 'cp0122']


def save_to_excel(users_data):
    # Generar un DataFrame a partir de los datos de usuario
    df = pd.DataFrame(users_data)

    # Verificar si el archivo Excel existe
    filename = 'usuarios.xlsx'
    if not os.path.exists(filename):
        # Si no existe, crear un nuevo archivo Excel
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Hoja1', index=False)
            # Agregar las columnas para los resultados de las pruebas
            for test_name in tests_names:
                writer.sheets['Hoja1'].cell(row=1, column=len(df.columns) + 1, value=test_name)
    else:
        # Si el archivo existe, abrirlo en modo de adición
        with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
            # Obtener el nombre de la nueva hoja
            sheet_name = f'Hoja{len(writer.book.sheetnames) + 1}'
            # Guardar el DataFrame en una nueva hoja
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            # Agregar las columnas para los resultados de las pruebas
            for test_name in tests_names:
                writer.sheets[sheet_name].cell(row=1, column=len(df.columns) + 1, value=test_name)

    # Ejecutar las pruebas para cada usuario y registrar los resultados en el archivo Excel
    with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
        for i, row in df.iterrows():
            print(f"User #{i}")
            name, username, email, password = row['name'], row['username'], row['email'], row['password']
            for test_name in tests_names:
                # Ejecutar la prueba correspondiente
                test_result = globals()[f'{test_name.lower()}'](name, username, email, password)

if __name__ == "__main__":
    num_users = int(input("Ingrese el número de usuarios a generar: "))
    users_data = generate_user_data(num_users)
    save_to_excel(users_data)
